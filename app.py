

from openpyxl.reader.excel import load_workbook


def automate_spreadsheet():
    load_xl_file = load_workbook("Python Automation.xlsx")
    access_sheet = load_xl_file["Sheet1"]

    for row in range (3, access_sheet.max_row + 1):
        access_cell = access_sheet.cell(row, 4)
        corrected_price = access_cell.value * 0.9
        corrected_price_cell = access_sheet.cell(row, 5)
        corrected_price_cell.value = corrected_price

        load_xl_file.save('Corrected Spreadsheet.xlsx')